import pandas as pd
import json

file_path = r"C:\Users\arthu\.gemini\antigravity\scratch\aroom_health_analytics\Catalogação - Custo materia prima e insumos Aroom.xlsx"
csv_path = r"C:\Users\arthu\.gemini\antigravity\scratch\aroom_health_analytics\custos_limpos.csv"

try:
    # Just try reading the second sheet (index 1) since the first one was empty
    df = pd.read_excel(file_path, sheet_name=1)
    print("Columns in sheet 1:", df.columns.tolist()[:10])
    
    # Save the head for debugging
    print("Head:\n", df.head(3))
    
    # If the second sheet fails or is empty, we can just read all sheets using openpyxl directly
except Exception as e:
    print("Error reading sheet 1:", e)
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheet names:", wb.sheetnames)
    for s in wb.sheetnames:
        sheet = wb[s]
        print(f"Sheet {s} has {sheet.max_row} rows and {sheet.max_column} cols.")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i < 3:
                print(row)
            else:
                break
